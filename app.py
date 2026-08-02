from flask import Flask, render_template, Blueprint, request,jsonify, redirect, url_for, session
from werkzeug.security import generate_password_hash, check_password_hash
from models.user import db, User
from datetime import datetime, date, timedelta 
import calendar
import os
from models.transaction import Transaction
from models.budget import Budget
from models.investment import Investment
from openpyxl import Workbook, load_workbook
from flask import send_file
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
import io
from flask import flash
from werkzeug.utils import secure_filename
from models.loan import Loan
from routes.assistant import assistant_bp
from ai.ai_engine import ask_ai
from models.wallet import Wallet
from models.savings_goal import SavingsGoal
from flask_mail import Mail, Message
import random
from models.loan import Loan
from models.savings_goal import SavingsGoal
from sqlalchemy import func

app = Flask(__name__)
app.secret_key = "change_this_to_a_long_random_secret_key"

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

os.makedirs("database", exist_ok=True)

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///database/finance.db"
app.config["SQLALCHEMY_DATABASE_URI"] = (
    "sqlite:///" + os.path.join(BASE_DIR, "database", "finance.db")
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = "static/uploads"
app.config["MAIL_SERVER"] = "smtp-relay.brevo.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USE_SSL"] = False
app.config["MAIL_TIMEOUT"] = 10

app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD")
mail = Mail(app)
def send_otp(email, otp):

    import smtplib

    try:
        print("Testing SMTP...")
        print("Connecting to smtp-relay.brevo.com:587")
        server = smtplib.SMTP("smtp-relay.brevo.com", 587, timeout=10)
        server.starttls()
        print("TLS started")
        server.login(app.config["MAIL_USERNAME"], app.config["MAIL_PASSWORD"])

        print("✅ SMTP LOGIN SUCCESS")

        server.quit()

        return True

    except Exception as e:
        print("❌ SMTP ERROR:", e)
        return False
db.init_app(app)

with app.app_context():
    db.create_all()

@app.route("/")
def home():
    return redirect(url_for("login"))
@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":

        fullname = request.form["fullname"].strip()
        email = request.form["email"].strip().lower()

        existing_user = User.query.filter_by(email=email).first()

        if existing_user:
            flash("Email already exists!")
            return redirect(url_for("register"))

        otp = str(random.randint(100000, 999999))

        session["register_name"] = fullname
        session["register_email"] = email
        session["register_otp"] = otp

        print("OTP:", otp)

        if not send_otp(email, otp):
            flash("Unable to send OTP email. Please try again later.")
            return redirect(url_for("register"))

        return redirect(url_for("verify_register_otp"))

    return render_template("register.html")
@app.route("/verify-register-otp", methods=["GET", "POST"])
def verify_register_otp():

    if "register_email" not in session:
        return redirect(url_for("register"))

    if request.method == "POST":

        entered_otp = request.form["otp"].strip()

        if entered_otp == session.get("register_otp"):

            session["register_verified"] = True

            return redirect(url_for("create_password"))

        flash("Invalid OTP.")

    return render_template("verify_register_otp.html")
@app.route("/create-password", methods=["GET", "POST"])
def create_password():

    if not session.get("register_verified"):
        return redirect(url_for("register"))

    if request.method == "POST":

        password = request.form["password"]
        confirm = request.form["confirm_password"]

        if password != confirm:
            flash("Passwords do not match.")
            return redirect(url_for("create_password"))

        hashed_password = generate_password_hash(password)

        new_user = User(
            full_name=session["register_name"],
            email=session["register_email"],
            password=hashed_password
        )

        db.session.add(new_user)
        db.session.commit()

        wallets = [
            Wallet(user_id=new_user.id, name="Cash"),
            Wallet(user_id=new_user.id, name="SBI"),
            Wallet(user_id=new_user.id, name="HDFC"),
            Wallet(user_id=new_user.id, name="PhonePe"),
            Wallet(user_id=new_user.id, name="Google Pay")
        ]

        db.session.add_all(wallets)
        db.session.commit()

        session["user_id"] = new_user.id

        session.pop("register_name", None)
        session.pop("register_email", None)
        session.pop("register_otp", None)
        session.pop("register_verified", None)

        flash("Account created successfully!")

        return redirect(url_for("dashboard"))

    return render_template("create_password.html")
@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        email = request.form["email"].strip().lower()
        password = request.form["password"]

        user = User.query.filter_by(email=email).first()

        # Account does not exist
        if not user:
            flash("Account not found. Please create an account first.")
            return redirect(url_for("register"))

        # Wrong password
        if not check_password_hash(user.password, password):
            flash("Incorrect password.")
            return redirect(url_for("login"))

        # Blocked account
        if user.is_blocked:
            flash("Your account has been blocked by the administrator.")
            return redirect(url_for("login"))

        # Login success
        session["user_id"] = user.id
        session["is_admin"] = user.is_admin

        return redirect(url_for("dashboard"))

    return render_template("login.html")
@app.route("/admin")
def admin():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    total_users = User.query.count()
    total_transactions = Transaction.query.count()
    total_wallets = Wallet.query.count()
    total_investments = Investment.query.count()
    total_loans = Loan.query.count()
    total_goals = SavingsGoal.query.count()

    total_income = db.session.query(
        func.sum(Transaction.amount)
    ).filter(
        Transaction.type == "Income"
    ).scalar() or 0

    total_expense = db.session.query(
        func.sum(Transaction.amount)
    ).filter(
        Transaction.type == "Expense"
    ).scalar() or 0

    return render_template(
        "admin.html",
        total_users=total_users,
        total_transactions=total_transactions,
        total_wallets=total_wallets,
        total_investments=total_investments,
        total_loans=total_loans,
        total_goals=total_goals,
        total_income=total_income,
        total_expense=total_expense
    )
@app.route("/admin/users")
def admin_users():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    users = User.query.all()

    return render_template("admin_users.html", users=users)
@app.route("/admin/delete-user/<int:id>")
def admin_delete_user(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    user = User.query.get_or_404(id)

    if user.is_admin:
        flash("Admin account cannot be deleted.")
        return redirect(url_for("admin_users"))

    db.session.delete(user)
    db.session.commit()

    flash("User deleted successfully.")

    return redirect(url_for("admin_users"))
@app.route("/admin/transactions")
def admin_transactions():
    income_count = Transaction.query.filter_by(type="Income").count()

    expense_count = Transaction.query.filter_by(type="Expense").count()

    total_count = Transaction.query.count()

    today_count = Transaction.query.filter_by(
         date=date.today()
    ).count()

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    transactions = (
        db.session.query(Transaction, User)
        .join(User, Transaction.user_id == User.id)
        .order_by(Transaction.date.desc())
        .all()
    )

    return render_template(
        "admin_transactions.html",
        transactions=transactions,
        total_count=total_count,
        income_count=income_count,
        expense_count=expense_count,
        today_count=today_count
    )
@app.route("/admin/delete-transaction/<int:id>")
def admin_delete_transaction(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    transaction = Transaction.query.get_or_404(id)

    wallet = Wallet.query.get(transaction.wallet_id)

    if wallet:

        if transaction.type == "Income":
            wallet.balance -= transaction.amount

        else:
            wallet.balance += transaction.amount

    db.session.delete(transaction)

    db.session.commit()

    flash("Transaction deleted successfully!")

    return redirect(url_for("admin_transactions"))

@app.route("/admin/edit-transaction/<int:id>", methods=["GET", "POST"])
def admin_edit_transaction(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    transaction = Transaction.query.get_or_404(id)

    wallets = Wallet.query.filter_by(
        user_id=transaction.user_id
    ).all()

    if request.method == "POST":

        old_wallet = Wallet.query.get(transaction.wallet_id)

        if old_wallet:

            if transaction.type == "Income":
                old_wallet.balance -= transaction.amount
            else:
                old_wallet.balance += transaction.amount

        transaction.wallet_id = int(request.form["wallet_id"])
        transaction.type = request.form["type"]
        transaction.category = request.form["category"]
        transaction.amount = float(request.form["amount"])
        transaction.note = request.form["note"]

        transaction.date = datetime.strptime(
            request.form["date"],
            "%Y-%m-%d"
        ).date()

        new_wallet = Wallet.query.get(transaction.wallet_id)

        if new_wallet:

            if transaction.type == "Income":
                new_wallet.balance += transaction.amount
            else:
                new_wallet.balance -= transaction.amount

        db.session.commit()

        flash("Transaction updated successfully!")

        return redirect(url_for("admin_transactions"))

    return render_template(
        "admin_edit_transaction.html",
        transaction=transaction,
        wallets=wallets
    )
@app.route("/admin/wallets")
def admin_wallets():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    wallets = db.session.query(Wallet, User).join(
        User, Wallet.user_id == User.id
    ).all()

    return render_template(
        "admin_wallets.html",
        wallets=wallets
    )
@app.route("/admin/wallet/<int:id>")
def admin_view_wallet(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    wallet = Wallet.query.get_or_404(id)

    user = User.query.get(wallet.user_id)

    return render_template(
        "admin_view_wallet.html",
        wallet=wallet,
        user=user
    )
@app.route("/admin/edit-wallet/<int:id>", methods=["GET", "POST"])
def admin_edit_wallet(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    wallet = Wallet.query.get_or_404(id)

    if request.method == "POST":

        wallet.name = request.form["name"]
        wallet.balance = float(request.form["balance"])

        db.session.commit()

        flash("Wallet updated successfully!")

        return redirect(url_for("admin_wallets"))

    return render_template(
        "admin_edit_wallet.html",
        wallet=wallet
    )
@app.route("/admin/delete-wallet/<int:id>")
def admin_delete_wallet(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    wallet = Wallet.query.get_or_404(id)

    db.session.delete(wallet)
    db.session.commit()

    flash("Wallet deleted successfully!")

    return redirect(url_for("admin_wallets"))
@app.route("/admin/wallet-history/<int:id>")
def admin_wallet_history(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    wallet = Wallet.query.get_or_404(id)

    transactions = Transaction.query.filter_by(
        wallet_id=wallet.id
    ).order_by(Transaction.date.desc()).all()

    return render_template(
        "admin_wallet_history.html",
        wallet=wallet,
        transactions=transactions
    )
@app.route("/admin/toggle-wallet/<int:id>")
def toggle_wallet(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    wallet = Wallet.query.get_or_404(id)

    wallet.is_active = not wallet.is_active

    db.session.commit()

    flash("Wallet status updated successfully!")

    return redirect(url_for("admin_wallets"))
@app.route("/admin/investments")
def admin_investments():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    investments = db.session.query(Investment, User).join(
        User, Investment.user_id == User.id
    ).all()

    return render_template(
        "admin_investments.html",
        investments=investments
    )
@app.route("/admin/investment/edit/<int:id>", methods=["GET", "POST"])
def admin_edit_investment(id):
    pass

@app.route("/admin/investment/delete/<int:id>")
def admin_delete_investment(id):
    pass
@app.route("/admin/loans")
def admin_loans():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    loans = db.session.query(Loan, User).join(
        User, Loan.user_id == User.id
    ).all()

    return render_template(
        "admin_loans.html",
        loans=loans
    )
@app.route("/admin/loan/edit/<int:id>", methods=["GET", "POST"])
def admin_edit_loan(id):
    pass

@app.route("/admin/loan/delete/<int:id>")
def admin_delete_loan(id):
    pass
@app.route("/admin/goals")
def admin_goals():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    goals = db.session.query(SavingsGoal, User).join(
        User, SavingsGoal.user_id == User.id
    ).all()

    return render_template(
        "admin_goals.html",
        goals=goals
    )
@app.route("/goal/add-money/<int:id>", methods=["GET", "POST"])
def add_money_to_goal(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    goal = SavingsGoal.query.get_or_404(id)

    if goal.user_id != session["user_id"]:
        return "Unauthorized"

    if request.method == "POST":

        amount = float(request.form["amount"])

        print("Amount entered:", amount)
        print("Before:", goal.saved_amount)

        goal.saved_amount += amount

        print("After:", goal.saved_amount)

        db.session.commit()

        print("Committed to database")
        flash("Money added successfully!")

        return redirect(url_for("savings_goals"))
    return render_template(
        "add_goal_money.html",
        goal=goal
    )
@app.route("/admin/goal/edit/<int:id>", methods=["GET", "POST"])
def admin_edit_goal(id):
    pass

@app.route("/admin/goal/delete/<int:id>")
def admin_delete_goal(id):
    pass
@app.route("/admin/toggle-block/<int:id>")
def toggle_block(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    user = User.query.get_or_404(id)

    if user.is_admin:
        flash("Admin account cannot be blocked.")
        return redirect(url_for("admin_users"))

    user.is_blocked = not user.is_blocked

    db.session.commit()

    if user.is_blocked:
        flash("User blocked successfully.")
    else:
        flash("User unblocked successfully.")

    return redirect(url_for("admin_users"))
@app.route("/admin/analytics")
def admin_analytics():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if not session.get("is_admin"):
        return "Unauthorized"

    total_users = User.query.count()
    total_transactions = Transaction.query.count()
    total_wallets = Wallet.query.count()
    total_investments = Investment.query.count()
    total_loans = Loan.query.count()
    total_goals = SavingsGoal.query.count()

    # 👇 ADD THIS HERE
    user_transaction_stats = (
        db.session.query(
            User.full_name,
            func.count(Transaction.id)
        )
        .join(Transaction, User.id == Transaction.user_id)
        .group_by(User.id)
        .all()
    )
    print(user_transaction_stats)
    return render_template(
        "admin_analytics.html",
        total_users=total_users,
        total_transactions=total_transactions,
        total_wallets=total_wallets,
        total_investments=total_investments,
        total_loans=total_loans,
        total_goals=total_goals,
        user_transaction_stats=user_transaction_stats
    )
@app.route("/profile")
def profile():

    if "user_id" not in session:
        return redirect(url_for("login"))

    user = User.query.get(session["user_id"])

    return render_template("profile.html", user=user)

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():

    if request.method == "POST":

        email = request.form["email"]

        user = User.query.filter_by(email=email).first()

        if not user:
            flash("Email not found.")
            return redirect(url_for("forgot_password"))

        otp = str(random.randint(100000, 999999))

        session["reset_email"] = email
        session["reset_otp"] = otp
        session["otp_expiry"] = (
            datetime.now() + timedelta(minutes=5)
        ).strftime("%Y-%m-%d %H:%M:%S")
        msg = Message(
            "Password Reset OTP",
            sender=app.config["MAIL_USERNAME"],
            recipients=[email]
        )

        msg.body = f"Your OTP is: {otp}"

        mail.send(msg)

        flash("OTP sent to your email.")

        return redirect(url_for("verify_otp"))

    return render_template("forgot_password.html")
@app.route("/logout")
def logout():

    session.clear()

    return redirect(url_for("login"))

@app.route("/verify-otp", methods=["GET", "POST"])
def verify_otp():

    if request.method == "POST":

        otp = request.form["otp"]

        expiry = datetime.strptime(
            session["otp_expiry"],
            "%Y-%m-%d %H:%M:%S"
        )

        if datetime.now() > expiry:
            flash("OTP has expired.")
            return redirect(url_for("forgot_password"))

        if otp == session.get("reset_otp"):
            return redirect(url_for("reset_password"))
        else:
            flash("Invalid OTP")
            return redirect(url_for("verify_otp"))

    return render_template("verify_otp.html")
@app.route("/resend-otp")
def resend_otp():

    if "reset_email" not in session:
        return redirect(url_for("forgot_password"))

    otp = str(random.randint(100000,999999))

    session["reset_otp"] = otp
    session["otp_expiry"] = (
        datetime.now() + timedelta(minutes=5)
    ).strftime("%Y-%m-%d %H:%M:%S")

    msg = Message(
        "Password Reset OTP",
        sender=app.config["MAIL_USERNAME"],
        recipients=[session["reset_email"]]
    )

    msg.body = f"Your new OTP is: {otp}"

    mail.send(msg)

    flash("New OTP sent.")

    return redirect(url_for("verify_otp"))
@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():

    if "reset_email" not in session:
        return redirect(url_for("forgot_password"))

    if request.method == "POST":

        password = request.form["password"]

        user = User.query.filter_by(
            email=session["reset_email"]
        ).first()

        user.password = generate_password_hash(password)

        db.session.commit()

        session.pop("reset_email", None)
        session.pop("reset_otp", None)
        session.pop("otp_expiry", None)
        flash("Password changed successfully!")

        return redirect(url_for("login"))

    return render_template("reset_password.html")
@app.route("/dashboard")
def dashboard():

    if "user_id" not in session:
        return redirect(url_for("login"))

    hour = datetime.now().hour

    if hour < 12:
        greeting = " Good Morning"
    elif hour < 17:
        greeting = "Good Afternoon"
    else:
        greeting = "Good Evening"

    user = User.query.get(session["user_id"])
    notifications = []
    wallets = Wallet.query.filter_by(user_id=user.id).all()

    search = request.args.get("search", "")
    type_filter = request.args.get("type", "")

    query = Transaction.query.filter_by(user_id=user.id)

    if search:
       query = query.filter(Transaction.category.contains(search))

    if type_filter:
       query = query.filter(Transaction.type == type_filter)

    transactions = query.order_by(Transaction.date.desc()).all()

    income = sum(t.amount for t in transactions if t.type == "Income")
    expense = sum(t.amount for t in transactions if t.type == "Expense")
    balance = income - expense
    savings = balance

    if income > 0:
       budget_used = round((expense / income) * 100)
    else:
       budget_used = 0

# Category totals
    food = sum(t.amount for t in transactions if t.type == "Expense" and t.category == "Food")
    fuel = sum(t.amount for t in transactions if t.type == "Expense" and t.category == "Fuel")
    shopping = sum(t.amount for t in transactions if t.type == "Expense" and t.category == "Shopping")
    travel = sum(t.amount for t in transactions if t.type == "Expense" and t.category == "Travel")
    other = sum(
       t.amount for t in transactions
       if t.type == "Expense"
       and t.category not in [
        "Food", "Fuel", "Shopping", "Travel",
        "Bills", "Medical", "Entertainment", "Education"
       ]
    )

# Budget checking
    budgets = Budget.query.filter_by(user_id=user.id).all()
    print("Budgets:", [(b.category, b.amount) for b in budgets])
    budget_warnings = []

    expense_map = {
    "Food": food,
    "Fuel": fuel,
    "Shopping": shopping,
    "Travel": travel,
    "Bills": sum(t.amount for t in transactions if t.type == "Expense" and t.category == "Bills"),
    "Medical": sum(t.amount for t in transactions if t.type == "Expense" and t.category == "Medical"),
    "Entertainment": sum(t.amount for t in transactions if t.type == "Expense" and t.category == "Entertainment"),
    "Education": sum(t.amount for t in transactions if t.type == "Expense" and t.category == "Education"),
    "Other": other
    }

    for b in budgets:
       spent = expense_map.get(b.category, 0)

       if spent > b.amount:
           budget_warnings.append({
            "category": b.category,
            "budget": b.amount,
            "spent": spent,
            "exceeded": spent - b.amount
      })
    for warning in budget_warnings:
        notifications.append({
        "icon": "🔴",
        "message": f"{warning['category']} budget exceeded by ₹{warning['exceeded']:.2f}"
    })
    print("Budgets:", [(b.category, b.amount) for b in budgets])
    print("Warnings:", budget_warnings)
    # ---------- AI INSIGHTS ----------
# ---------- AI INSIGHTS ----------

    ai_insights = []

    expense_categories = {
    "Food": food,
    "Fuel": fuel,
    "Shopping": shopping,
    "Travel": travel,
    "Other": other
    }

    highest_category = max(expense_categories, key=expense_categories.get)
    highest_amount = expense_categories[highest_category]

    if highest_amount > 0:
       ai_insights.append(
        f"🛍 Your highest spending category is {highest_category} (₹{highest_amount:.2f})."
       )

    if savings > 0:
       ai_insights.append(
        f"💰 Great! You saved ₹{savings:.2f} this month."
       )
    else:
       ai_insights.append(
        f"⚠️ Your expenses exceeded your income by ₹{abs(savings):.2f}."
       )

    if income > 0:
       saving_percent = (savings / income) * 100

       if saving_percent >= 20:
          ai_insights.append(
            "🎯 Excellent! You are saving more than 20% of your income."
          )
       elif saving_percent >= 10:
          ai_insights.append(
            "👍 Good job! Your savings are on track."
          )
       else:
          ai_insights.append(
            "💡 Try reducing unnecessary expenses to improve your savings."
          )
    # Financial Health Score
# Financial Health Score
    # Financial Health Score
    if income == 0 and expense == 0:
        health_score = 0
    else:
        health_score = 100

        if income > 0:
            savings_percent = ((income - expense) / income) * 100
        else:
            savings_percent = 0

        if savings_percent < 10:
            health_score -= 20
        elif savings_percent < 20:
            health_score -= 10

        # Budget Penalty
        health_score -= len(budget_warnings) * 5

        # Loan Penalty
        pending_loans = Loan.query.filter_by(
            user_id=user.id,
            status="Pending"
        ).count()

        health_score -= pending_loans * 5

    health_score = max(0, min(100, health_score))

    return render_template(
        "dashboard.html",
        greeting=greeting,
        user=user,
        transactions=transactions,
        income=income,
        expense=expense,
        balance=balance,
        savings=savings,
        budget_used=budget_used,
        budget_warnings=budget_warnings,
        food=food,
        fuel=fuel,
        shopping=shopping,
        health_score=health_score,
        travel=travel,
        other=other,
        ai_insights=ai_insights,
        wallets=wallets,
        notifications=notifications
    )

@app.route("/add-transaction", methods=["GET", "POST"])
def add_transaction():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":

        # Upload receipt
        receipt_name = None

        receipt = request.files.get("receipt")

        if receipt and receipt.filename != "":
            filename = secure_filename(receipt.filename)
            receipt.save(
                os.path.join(app.config["UPLOAD_FOLDER"], filename)
            )
            receipt_name = filename
        wallet = Wallet.query.get(int(request.form["wallet_id"]))
        # Create transaction
        transaction = Transaction(
            user_id=session["user_id"],
            wallet_id=wallet.id,
            type=request.form["type"],
            category=request.form["category"],
            amount=float(request.form["amount"]),
            note=request.form["note"],
            date=datetime.strptime(
                request.form["date"],
                "%Y-%m-%d"
            ).date(),
            receipt=receipt_name,
            recurring=True if request.form.get("recurring") else False,
            frequency=request.form.get("frequency"),
            next_due=datetime.strptime(
            request.form["date"],
            "%Y-%m-%d"
            ).date()
        )

        db.session.add(transaction)
        if transaction.type == "Income":
            wallet.balance += transaction.amount
        else:
            wallet.balance -= transaction.amount
        db.session.commit()

        flash("Transaction added successfully!")

        return redirect(url_for("dashboard"))

    wallets = Wallet.query.filter_by(user_id=session["user_id"]).all()

    print(wallets)

    return render_template(
    "add_transaction.html",
    wallets=wallets
    )

@app.route("/edit-transaction/<int:id>", methods=["GET", "POST"])
def edit_transaction(id):

    transaction = Transaction.query.get_or_404(id)

    if transaction.user_id != session["user_id"]:
        return "Unauthorized"

    wallets = Wallet.query.filter_by(user_id=session["user_id"]).all()

    if request.method == "POST":

        # Restore old wallet balance
        old_wallet = Wallet.query.get(transaction.wallet_id)

        if old_wallet:
            if transaction.type == "Income":
                old_wallet.balance -= transaction.amount
            else:
                old_wallet.balance += transaction.amount

        # Update transaction
        wallet_id = request.form.get("wallet_id")

        if wallet_id:
            transaction.wallet_id = int(wallet_id)
        else:
            transaction.wallet_id = None
        transaction.type = request.form["type"]
        transaction.category = request.form["category"]
        transaction.amount = float(request.form["amount"])
        transaction.note = request.form["note"]
        transaction.date = datetime.strptime(
            request.form["date"],
            "%Y-%m-%d"
        ).date()

        # Apply to new wallet
        new_wallet = Wallet.query.get(transaction.wallet_id)

        if new_wallet:
            if transaction.type == "Income":
                new_wallet.balance += transaction.amount
            else:
                new_wallet.balance -= transaction.amount

        db.session.commit()

        flash("Transaction updated successfully!")

        return redirect(url_for("dashboard"))

    return render_template(
        "edit_transaction.html",
        transaction=transaction,
        wallets=wallets
    )
@app.route("/delete-transaction/<int:id>")
def delete_transaction(id):

    transaction = Transaction.query.get_or_404(id)

    if transaction.user_id != session["user_id"]:
        return "Unauthorized"

    wallet = Wallet.query.get(transaction.wallet_id)

    if wallet:

        if transaction.type == "Income":
            wallet.balance -= transaction.amount

        else:
            wallet.balance += transaction.amount

    db.session.delete(transaction)

    db.session.commit()

    return redirect(url_for("dashboard"))
@app.route("/report")
def report():

    if "user_id" not in session:
        return redirect(url_for("login"))

    user = User.query.get(session["user_id"])

    transactions = Transaction.query.filter_by(user_id=user.id).all()

    income = sum(t.amount for t in transactions if t.type == "Income")
    expense = sum(t.amount for t in transactions if t.type == "Expense")
    savings = income - expense

    return render_template(
        "report.html",
        user=user,
        income=income,
        expense=expense,
        savings=savings,
        transactions=transactions
    )
@app.route("/budget", methods=["GET", "POST"])
def budget():

    if "user_id" not in session:
        return redirect(url_for("login"))

    user_id = session["user_id"]

    if request.method == "POST":
        print("POST RECEIVED")
        print(request.form)

        category = request.form["category"]
        amount = float(request.form["amount"])

        budget = Budget.query.filter_by(
            user_id=user_id,
            category=category
        ).first()

        if budget:
            budget.amount = amount
        else:
            budget = Budget(
                user_id=user_id,
                category=category,
                amount=amount
            )
            db.session.add(budget)

        db.session.commit()
        flash("✅ Budget saved successfully!")
        return redirect(url_for("budget"))

    budgets = Budget.query.filter_by(
        user_id=user_id
    ).all()

    budget_data=[]

    for budget in budgets:

        spent = db.session.query(
            db.func.sum(Transaction.amount)
        ).filter(
            Transaction.user_id==user_id,
            Transaction.type=="Expense",
            Transaction.category==budget.category
        ).scalar() or 0

        remaining=max(
            budget.amount-spent,
            0
        )

        percent=0

        if budget.amount>0:
            percent=(spent/budget.amount)*100

        budget_data.append({
            "id": budget.id,
            "category":budget.category,

            "budget":budget.amount,

            "spent":spent,

            "remaining":remaining,

            "percent":percent

        })

    return render_template(
        "budget.html",
        budget_data=budget_data
    )
@app.route("/reset-budget")
def reset_budget():

    if "user_id" not in session:
        return redirect(url_for("login"))

    Budget.query.filter_by(user_id=session["user_id"]).delete()

    db.session.commit()

    return redirect(url_for("budget"))
@app.route("/delete-budget/<int:id>")
def delete_budget(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    budget = Budget.query.get_or_404(id)

    if budget.user_id != session["user_id"]:
        return "Unauthorized"

    db.session.delete(budget)
    db.session.commit()

    flash("🗑 Budget deleted successfully!")

    return redirect(url_for("budget"))
@app.route("/export-excel")
def export_excel():

    if "user_id" not in session:
        return redirect(url_for("login"))

    transactions = Transaction.query.filter_by(
        user_id=session["user_id"]
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append([
        "Date",
        "Category",
        "Type",
        "Amount",
        "Note"
    ])

    for t in transactions:
        ws.append([
            str(t.date),
            t.category,
            t.type,
            t.amount,
            t.note
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="transactions.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/export-pdf")
def export_pdf():

    if "user_id" not in session:
        return redirect(url_for("login"))

    transactions = Transaction.query.filter_by(
        user_id=session["user_id"]
    ).all()

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(buffer)

    elements = []

    styles = getSampleStyleSheet()

    elements.append(Paragraph("Personal Finance Report", styles["Title"]))

    data = [["Date", "Category", "Type", "Amount (₹)", "Note"]]

    for t in transactions:
        data.append([
            str(t.date),
            t.category,
            t.type,
            f"₹{t.amount}",
            t.note or ""
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.blue),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BACKGROUND", (0,1), (-1,-1), colors.beige),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("BOTTOMPADDING", (0,0), (-1,0), 10),
    ]))

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="Finance_Report.pdf",
        mimetype="application/pdf"
    )
@app.route("/import-excel", methods=["GET", "POST"])
def import_excel():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":

        file = request.files.get("excel_file")

        if not file:
            flash("Please select an Excel file.")
            return redirect(request.url)

        workbook = load_workbook(file)
        sheet = workbook.active

        headers = [
            str(cell.value).strip().lower()
            if cell.value else ""
            for cell in sheet[1]
        ]

        print(headers)

        imported = 0
        skipped = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):

            data = dict(zip(headers, row))

            def get_value(*names):
                for name in names:
                    if name.lower() in data:
                        return data[name.lower()]
                return None

            date = get_value("Date")

            type_ = get_value(
                "Type",
                "Type (Income/Expense/Transfer)"
            )

            category = get_value(
                "Category",
                "Budget Category"
            )

            amount = get_value("Amount")

            wallet = get_value(
                "Wallet/Account",
                "Wallet"
            )

            note = get_value(
                "Notes",
                "Note",
                "Description"
            )

            if note is None:
                note = ""

            try:

                if (
                    not date
                    or not type_
                    or not category
                    or amount is None
                ):
                    skipped += 1
                    continue

                if isinstance(date, datetime):
                    date = date.date()

                elif isinstance(date, str):

                    parsed = None

                    for fmt in (
                        "%Y-%m-%d",
                        "%d-%m-%Y",
                        "%d/%m/%Y",
                        "%m/%d/%Y"
                    ):
                        try:
                            parsed = datetime.strptime(
                                date,
                                fmt
                            ).date()
                            break
                        except:
                            pass

                    if parsed is None:
                        skipped += 1
                        continue

                    date = parsed

                wallet_id = None

                if wallet:

                    wallet_obj = Wallet.query.filter_by(
                        user_id=session["user_id"],
                        name=str(wallet).strip()
                    ).first()

                    if wallet_obj:
                        wallet_id = wallet_obj.id

                transaction = Transaction(
                    user_id=session["user_id"],
                    date=date,
                    type=str(type_).strip(),
                    category=str(category).strip(),
                    amount=float(amount),
                    note=str(note),
                    wallet_id=wallet_id
                )

                db.session.add(transaction)

                imported += 1

            except Exception as e:
                print(e)
                skipped += 1

        db.session.commit()

        flash(
            f"Imported {imported} transactions. Skipped {skipped} rows."
        )

        return redirect(url_for("dashboard"))

    return render_template("import_excel.html")
@app.route("/analytics")
def analytics():

    if "user_id" not in session:
        return redirect(url_for("login"))

    user_id = session["user_id"]

    transactions = Transaction.query.filter_by(user_id=user_id).all()

    total_income = 0
    total_expense = 0

    category_data = {}

    monthly_data = {}

    for t in transactions:

        month = t.date.strftime("%b %Y")

        monthly_data.setdefault(month, {"income": 0, "expense": 0})

        if t.type == "Income":

            total_income += t.amount
            monthly_data[month]["income"] += t.amount

        else:

            total_expense += t.amount
            monthly_data[month]["expense"] += t.amount

            category_data[t.category] = (
                category_data.get(t.category, 0)
                + t.amount
            )

    savings = total_income - total_expense
    print(monthly_data)
    return render_template(
        "analytics.html",
        income=total_income,
        expense=total_expense,
        savings=savings,
        category_data=category_data,
        monthly_data=monthly_data
    )
@app.route("/savings-goals", methods=["GET", "POST"])
def savings_goals():

    if "user_id" not in session:
        return redirect(url_for("login"))

    user_id = session["user_id"]

    if request.method == "POST":

        goal = SavingsGoal(
             user_id=session["user_id"],
             title=request.form["title"],
        target_amount=float(request.form["target_amount"]),
             saved_amount=0
        )

        db.session.add(goal)

        db.session.commit()

        return redirect(
            url_for("savings_goals")
        )

    goals = SavingsGoal.query.filter_by(
        user_id=user_id
    ).all()

    income = db.session.query(
        db.func.sum(Transaction.amount)
    ).filter_by(
        user_id=user_id,
        type="Income"
    ).scalar() or 0

    expense = db.session.query(
        db.func.sum(Transaction.amount)
    ).filter_by(
        user_id=user_id,
        type="Expense"
    ).scalar() or 0

    savings = income - expense
    for goal in goals:
        if savings >= goal.target_amount:
            notifications.append({
            "icon": "🎯",
            "message": f"You achieved your goal: {goal.title}"
        })
    return render_template(

        "savings_goals.html",

        goals=goals,

        savings=savings

    )
@app.route("/delete-goal/<int:id>")
def delete_goal(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    goal = SavingsGoal.query.get_or_404(id)

    if goal.user_id != session["user_id"]:
        return "Unauthorized"

    db.session.delete(goal)
    db.session.commit()

    return redirect(url_for("savings_goals"))
@app.route("/edit-goal/<int:id>", methods=["GET", "POST"])
def edit_goal(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    goal = SavingsGoal.query.get_or_404(id)

    if goal.user_id != session["user_id"]:
        return "Unauthorized"

    if request.method == "POST":

        goal.title = request.form["title"]

        goal.target_amount = float(
            request.form["target_amount"]
        )

        db.session.commit()

        return redirect(url_for("savings_goals"))

    return render_template(
        "edit_goal.html",
        goal=goal
    )
@app.route("/investments")
def investments():

    if "user_id" not in session:
        return redirect(url_for("login"))

    investments = Investment.query.filter_by(
        user_id=session["user_id"]
    ).all()

    total_invested = sum(i.invested_amount for i in investments)
    current_value = sum(i.current_value for i in investments)
    profit_loss = current_value - total_invested

    return render_template(
        "investments.html",
        investments=investments,
        total_invested=total_invested,
        current_value=current_value,
        profit_loss=profit_loss
    )
@app.route("/add-investment", methods=["GET", "POST"])
def add_investment():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":

        investment = Investment(
            user_id=session["user_id"],
            name=request.form["name"],
            type=request.form["type"],
            invested_amount=float(request.form["invested_amount"]),
            current_value=float(request.form["current_value"]),
            purchase_date=datetime.strptime(
                request.form["purchase_date"],
                "%Y-%m-%d"
            ).date()
        )

        db.session.add(investment)
        db.session.commit()

        flash("Investment added successfully!")

        return redirect(url_for("dashboard"))

    return render_template("add_investment.html")
@app.route("/add-loan", methods=["GET", "POST"])
def add_loan():

    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":

        loan = Loan(
            user_id=session["user_id"],
            loan_name=request.form["loan_name"],
            loan_type=request.form["loan_type"],
            loan_amount=float(request.form["loan_amount"]),
            interest_rate=float(request.form["interest_rate"]),
            emi_amount=float(request.form["emi_amount"]),
            outstanding=float(request.form["outstanding"]),
            due_date=datetime.strptime(
                request.form["due_date"],
                "%Y-%m-%d"
            ).date(),
            status="Pending"
        )

        db.session.add(loan)
        db.session.commit()

        flash("Loan added successfully!")

        return redirect(url_for("loans"))

    return render_template("add_loan.html")
@app.route("/loans")
def loans():

    if "user_id" not in session:
        return redirect(url_for("login"))

    loans = Loan.query.filter_by(user_id=session["user_id"]).all()

    return render_template(
        "loans.html",
        loans=loans
    )
@app.route("/edit-investment/<int:id>", methods=["GET", "POST"])
def edit_investment(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    investment = Investment.query.get_or_404(id)

    if investment.user_id != session["user_id"]:
        return "Unauthorized"

    if request.method == "POST":

        investment.name = request.form["name"]
        investment.type = request.form["type"]
        investment.invested_amount = float(request.form["invested_amount"])
        investment.current_value = float(request.form["current_value"])
        investment.purchase_date = datetime.strptime(
            request.form["purchase_date"],
            "%Y-%m-%d"
        ).date()

        db.session.commit()

        flash("Investment updated successfully!")

        return redirect(url_for("investments"))

    return render_template(
        "edit_investment.html",
        investment=investment
    )
@app.route("/mark-paid/<int:id>")
def mark_paid(id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    loan = Loan.query.get_or_404(id)

    loan.status = "Paid"

    db.session.commit()

    flash("Loan marked as paid!")

    return redirect(url_for("loans"))
@app.route("/edit-loan/<int:id>", methods=["GET", "POST"])
def edit_loan(id):
    loan = Loan.query.get_or_404(id)

    if request.method == "POST":
        loan.loan_name = request.form["loan_name"]
        loan.loan_type = request.form["loan_type"]
        loan.loan_amount = float(request.form["loan_amount"])
        loan.emi_amount = float(request.form["emi_amount"])
        loan.outstanding = float(request.form["outstanding"])
        loan.due_date = datetime.strptime(
            request.form["due_date"], "%Y-%m-%d"
        ).date()
        loan.status = request.form["status"]

        db.session.commit()
        flash("Loan updated successfully!")

        return redirect(url_for("loans"))

    return render_template("edit_loan.html", loan=loan)
@app.route("/delete-loan/<int:id>")
def delete_loan(id):
    loan = Loan.query.get_or_404(id)

    db.session.delete(loan)
    db.session.commit()

    flash("Loan deleted successfully!")

    return redirect(url_for("loans"))
@app.route("/delete-investment/<int:id>")
def delete_investment():

    if "user_id" not in session:
        return redirect(url_for("login"))

    investment = Investment.query.get_or_404(id)

    if investment.user_id != session["user_id"]:
        return "Unauthorized"
 
    db.session.delete(investment)
    db.session.commit()

    flash("Investment deleted successfully!")

    return redirect(url_for("investments"))
@app.route("/calendar")
def calendar_view():

    if "user_id" not in session:
        return redirect(url_for("login"))

    user_id = session["user_id"]

    today = date.today()

    year = request.args.get("year", default=today.year, type=int)
    month = request.args.get("month", default=today.month, type=int)

    cal = calendar.monthcalendar(year, month)

    transactions = Transaction.query.filter(
        Transaction.user_id == user_id
    ).all()

    daily_transactions = {}

    for t in transactions:

        if t.date.year == year and t.date.month == month:

            day = t.date.day

            if day not in daily_transactions:
                daily_transactions[day] = {
                    "income": 0,
                    "expense": 0
                }

            if t.type == "Income":
                daily_transactions[day]["income"] += t.amount
            else:
                daily_transactions[day]["expense"] += t.amount

    month_name = calendar.month_name[month]

    return render_template(
        "calendar.html",
        calendar_days=cal,
        daily_transactions=daily_transactions,
        month=month,
        year=year,
        month_name=month_name
    )
@app.route("/calendar/<int:year>/<int:month>/<int:day>")
def day_transactions(year, month, day):

    if "user_id" not in session:
        return redirect(url_for("login"))

    user_id = session["user_id"]

    selected_date = date(year, month, day)

    transactions = Transaction.query.filter_by(
        user_id=user_id,
        date=selected_date
    ).order_by(Transaction.id.desc()).all()

    return render_template(
        "day_transactions.html",
        transactions=transactions,
        selected_date=selected_date
    )
app.register_blueprint(assistant_bp)
def generate_ai_reply(balance, income, expense, savings, loans, investments, question):

    return f"""
    Your balance is ₹{balance}.
    Your income is ₹{income}.
    Your expenses are ₹{expense}.
    Your savings are ₹{savings}.
    Your question: {question}
    """
@app.route("/transfer", methods=["GET", "POST"])
def transfer():

    if "user_id" not in session:
        return redirect(url_for("login"))

    wallets = Wallet.query.filter_by(
        user_id=session["user_id"]
    ).all()

    if request.method == "POST":

        from_wallet = Wallet.query.get(
            int(request.form["from_wallet"])
        )

        to_wallet = Wallet.query.get(
            int(request.form["to_wallet"])
        )

        amount = float(request.form["amount"])

        if from_wallet.id == to_wallet.id:
            flash("Choose different wallets.")
            return redirect(url_for("transfer"))

        if from_wallet.balance < amount:
            flash("Insufficient balance.")
            return redirect(url_for("transfer"))

        from_wallet.balance -= amount
        to_wallet.balance += amount

        db.session.commit()

        flash("Money transferred successfully!")

        return redirect(url_for("dashboard"))

    return render_template(
        "transfer.html",
        wallets=wallets
    )
@app.route("/insights")
def insights():

    if "user_id" not in session:
        return redirect(url_for("login"))

    transactions = Transaction.query.filter_by(
        user_id=session["user_id"]
    ).all()

    total_income = 0
    total_expense = 0

    category_total = {}

    for t in transactions:

        if t.type == "Income":
            total_income += t.amount

        else:
            total_expense += t.amount

            category_total[t.category] = (
                category_total.get(t.category, 0)
                + t.amount
            )

    highest_category = None

    if category_total:
        highest_category = max(
            category_total,
            key=category_total.get
        )

    return render_template(
        "insights.html",
        income=total_income,
        expense=total_expense,
        highest_category=highest_category,
        category_total=category_total
    )
@app.route("/clear-transactions")
def clear_transactions():

    if "user_id" not in session:
        return redirect(url_for("login"))

    user_id = session["user_id"]

    # Delete only this user's transactions
    Transaction.query.filter_by(user_id=user_id).delete()

    # Reset all wallet balances for this user
    wallets = Wallet.query.filter_by(user_id=user_id).all()

    for wallet in wallets:
        wallet.balance = 0

    db.session.commit()

    flash("All transactions cleared successfully!")

    return redirect(url_for("dashboard"))
if __name__ == "__main__":
    app.run(debug=True)
